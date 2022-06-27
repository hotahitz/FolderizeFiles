from asyncio.windows_events import NULL
import openpyxl
from pathlib import Path
import os
import shutil

xlsx_file = Path('eternity products.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file) 

# Read the active sheet:
sheet = wb_obj.active
#print(sheet["A2"].value)
uniqueSizeList = []
for row in sheet.iter_rows(max_row=235):
    if not(row[2].value in uniqueSizeList) and (not(row[2].value == "Size")):
        uniqueSizeList.append(row[2].value)

parentDirectory = "Output"
sourceDirectory = "Kajaria Eternity Ultima"

def extractFileCode(name):
    file = row[4].value
    file = file[:-5]
    if file[-1] == "_":
        file = file[:-1]
    
    return file

for size in uniqueSizeList:
    if not os.path.isdir(parentDirectory + "/" +size):
        os.mkdir(parentDirectory + "/" +size)
    for row in sheet.iter_rows(max_row=235):
        if row[2].value == size:

            destFolder = parentDirectory+ "/" + size + "/" + row[0].value

            if not os.path.isdir(destFolder):
                os.mkdir(destFolder)
            try:
                if row[4].value and (not row[4].value == ""):
                    shutil.copy(sourceDirectory+ "/"+ row[4].value, destFolder)
                    if((not row[5].value) or row[5].value == ""):
                        os.rename(destFolder+"/"+row[4].value, destFolder+"/"+row[0].value+".jpg")
                    else :
                        os.rename(destFolder+"/"+row[4].value, destFolder+"/"+row[0].value+" F1.jpg")

                if row[5].value and (not row[5].value == ""):
                    shutil.copy(sourceDirectory+ "/"+ row[5].value, destFolder)
                    os.rename(destFolder+"/"+row[5].value, destFolder+"/"+row[0].value+" F2.jpg")

                if row[6].value and (not row[6].value == ""):
                    shutil.copy(sourceDirectory+ "/"+ row[6].value, destFolder)
                    os.rename(destFolder+"/"+row[6].value, destFolder+"/"+row[0].value+" F3.jpg")

                if row[7].value and (not row[7].value == ""):
                    shutil.copy(sourceDirectory+ "/"+ row[7].value, destFolder)
                    os.rename(destFolder+"/"+row[7].value, destFolder+"/"+row[0].value+" F4.jpg")

                if row[8].value and (not row[8].value == ""):
                    shutil.copy(sourceDirectory+ "/"+ row[8].value, destFolder)
                    os.rename(destFolder+"/"+row[8].value, destFolder+"/"+row[0].value+" F5.jpg")
            except:
                print("File not found")
            